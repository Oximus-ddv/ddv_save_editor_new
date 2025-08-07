"""
Excel data loading service for DDV Save Editor
"""
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import logging
from datetime import datetime
from openpyxl import load_workbook

from ..models.game_item import GameItem, ItemCategory, GameDatabase


logger = logging.getLogger(__name__)


class ExcelDataService:
    """Service for loading game data from Excel files"""
    
    def __init__(self, excel_path: str = "Disney Dream Light ID List - Mainted by Rubyelf.xlsx"):
        self.excel_path = Path(excel_path)
        self.last_modified = None
        self._cached_database: Optional[GameDatabase] = None
        
    def load_game_database(self, force_reload: bool = False) -> GameDatabase:
        """Load or reload the game database from Excel"""
        try:
            # Check if file exists
            if not self.excel_path.exists():
                logger.error(f"Excel file not found: {self.excel_path}")
                return GameDatabase()
            
            # Check if we need to reload
            current_modified = self.excel_path.stat().st_mtime
            if not force_reload and self._cached_database and current_modified == self.last_modified:
                logger.info("Using cached database (file unchanged)")
                return self._cached_database
            
            logger.info(f"Loading game data from: {self.excel_path}")
            self.last_modified = current_modified
            
            # Load Excel file
            excel_data = pd.read_excel(self.excel_path, sheet_name=None, engine='openpyxl')
            
            # Create database
            database = GameDatabase(
                last_updated=datetime.now().isoformat(),
                source_file=str(self.excel_path)
            )
            
            # Process each worksheet
            for sheet_name, df in excel_data.items():
                # Check if this sheet has a Category column (like your DDV Excel file)
                if 'Category' in df.columns or 'category' in df.columns:
                    items = self._process_categorized_worksheet(df, sheet_name)
                    for item in items:
                        database.add_item(item)
                    logger.info(f"Loaded {len(items)} items from '{sheet_name}' sheet (categorized)")
                else:
                    # Fall back to sheet name detection
                    category = self._detect_category(sheet_name)
                    if category:
                        items = self._process_worksheet(df, category, sheet_name)
                        for item in items:
                            database.add_item(item)
                        logger.info(f"Loaded {len(items)} items from '{sheet_name}' sheet")
            
            self._cached_database = database
            logger.info(f"Database loaded successfully: {database.get_stats()}")
            return database
            
        except Exception as e:
            logger.error(f"Error loading Excel data: {e}")
            return GameDatabase()
    
    def _detect_category(self, sheet_name: str) -> Optional[ItemCategory]:
        """Detect item category from worksheet name"""
        name_lower = sheet_name.lower().strip()
        
        # Category mappings
        category_mappings = {
            ItemCategory.PETS: ['pets', 'pet', 'companions', 'critters', 'animals'],
            ItemCategory.CLOTHES: ['clothes', 'clothing', 'apparel', 'outfits', 'fashion'],
            ItemCategory.HOUSES: ['houses', 'house', 'buildings', 'homes', 'structures'],
            ItemCategory.NPC_SKINS: ['skins', 'npc', 'characters', 'character_skins', 'npcskins'],
            ItemCategory.FURNITURE: ['furniture', 'decor', 'decoration', 'furnishing'],
            ItemCategory.TOOLS: ['tools', 'equipment', 'implements'],
            ItemCategory.FOOD: ['food', 'meals', 'cooking', 'recipes'],
            ItemCategory.MATERIALS: ['materials', 'resources', 'crafting', 'ingredients']
        }
        
        for category, keywords in category_mappings.items():
            if any(keyword in name_lower for keyword in keywords):
                return category
        
        # Skip ignored sheets
        ignored_patterns = ['sheet1', 'sheet', 'template', 'example', 'readme', 'info', 'instructions']
        if any(pattern in name_lower for pattern in ignored_patterns):
            return None
        
        logger.warning(f"Unknown category for sheet: {sheet_name}")
        return None
    
    def _process_categorized_worksheet(self, df: pd.DataFrame, sheet_name: str) -> List[GameItem]:
        """Process a worksheet that has its own Category column"""
        items = []
        
        if df.empty:
            logger.warning(f"Empty worksheet: {sheet_name}")
            return items
        
        # Normalize column names
        df.columns = [col.strip().lower() if isinstance(col, str) else str(col) for col in df.columns]
        
        # Find required columns
        id_col = self._find_column(df, ['id', 'itemid', 'item_id', 'item id'])
        name_col = self._find_column(df, ['name', 'itemname', 'item_name', 'item name'])
        category_col = self._find_column(df, ['category', 'type', 'item_type'])
        
        if not id_col or not name_col or not category_col:
            logger.warning(f"Missing required columns in {sheet_name}. Found: ID={id_col}, Name={name_col}, Category={category_col}")
            return items
        
        logger.info(f"Processing {sheet_name} with columns: ID={id_col}, Name={name_col}, Category={category_col}")
        
        # Optional columns
        image_col = self._find_column(df, ['image', 'icon', 'picture', 'img'])
        notes_col = self._find_column(df, ['notes', 'description', 'desc'])
        
        # Get colored cell information
        colored_info = self._get_colored_cell_info(sheet_name)
        
        # Get problematic rows to filter out
        problematic_rows = self._get_problematic_rows(df, notes_col)
        # Add blue cells to problematic rows (should be completely hidden)
        problematic_rows.update(colored_info['blue_rows'])
        
        # Process each row
        processed_count = 0
        skipped_count = 0
        for idx, row in df.iterrows():
            try:
                # Skip problematic rows (red cells or bad content)
                if idx in problematic_rows:
                    skipped_count += 1
                    if skipped_count <= 5:  # Log first few skipped rows
                        logger.debug(f"Skipped problematic row {idx}: {row[name_col] if name_col else 'unknown'}")
                    continue
                
                # Get basic info
                try:
                    item_id = int(row[id_col]) if pd.notna(row[id_col]) else None
                    if item_id is None:
                        skipped_count += 1
                        continue
                except (ValueError, TypeError):
                    skipped_count += 1
                    continue
                
                name = str(row[name_col]) if pd.notna(row[name_col]) else "Unknown Item"
                category_str = str(row[category_col]) if pd.notna(row[category_col]) else ""
                
                # Map category string to ItemCategory
                category = self._map_category_string(category_str)
                if not category:
                    skipped_count += 1
                    if idx < 10:  # Debug first 10 skipped items
                        logger.debug(f"Skipped row {idx}: category '{category_str}' not mapped")
                    continue
                
                # Optional fields
                image_path = str(row[image_col]) if image_col and pd.notna(row[image_col]) else None
                description = str(row[notes_col]) if notes_col and pd.notna(row[notes_col]) else None
                
                # Check if this is a yellow item (limit to 1)
                max_quantity = 1 if idx in colored_info['yellow_rows'] else None
                
                # Create item
                item = GameItem(
                    id=item_id,
                    name=name,
                    category=category,
                    description=description,
                    image_path=image_path,
                    max_quantity=max_quantity
                )
                items.append(item)
                processed_count += 1
                
            except Exception as e:
                logger.debug(f"Error processing row {idx} in {sheet_name}: {e}")
                skipped_count += 1
                continue
        
        logger.info(f"Processed {processed_count} items from categorized worksheet '{sheet_name}' (skipped {skipped_count})")
        return items
    
    def _get_colored_cell_info(self, sheet_name: str) -> dict:
        """Get information about colored cells in the Excel file"""
        try:
            wb = load_workbook(self.excel_path, read_only=True)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            
            colored_info = {
                'red_rows': set(),      # Problematic items (already handled by text)
                'blue_rows': set(),     # Hide completely
                'yellow_rows': set(),   # Limit to 1 item
            }
            
            # Color definitions
            red_color = 'FFFF0000'      # Red - problematic
            yellow_color = 'FFFBBC04'   # Yellow - limit to 1
            blue_color = 'FF00FFFF'     # Blue - hide completely
            
            # Check a reasonable range for performance (first 500 rows to catch more yellow items)
            max_row = min(500, ws.max_row or 100)
            
            for row_num in range(1, max_row):
                for col_num in range(1, 13):  # Check first 12 columns (A-L)
                    try:
                        cell = ws.cell(row=row_num, column=col_num)
                        if cell.fill and cell.fill.start_color:
                            color = str(cell.fill.start_color.index)
                            
                            if color == red_color:
                                colored_info['red_rows'].add(row_num - 1)  # Convert to 0-based
                            elif color == yellow_color:
                                colored_info['yellow_rows'].add(row_num - 1)
                            elif color == blue_color:
                                colored_info['blue_rows'].add(row_num - 1)
                    except:
                        continue
            
            wb.close()
            logger.info(f"Color analysis: {len(colored_info['red_rows'])} red, {len(colored_info['yellow_rows'])} yellow, {len(colored_info['blue_rows'])} blue rows")
            return colored_info
            
        except Exception as e:
            logger.warning(f"Could not analyze colored cells: {e}")
            return {'red_rows': set(), 'blue_rows': set(), 'yellow_rows': set()}
    
    def _get_problematic_rows(self, df: pd.DataFrame, notes_col: Optional[str]) -> set:
        """Get set of row indices that should be filtered out (text-based + blue cells)"""
        problematic_rows = set()
        
        # Check for problematic content indicators
        problematic_indicators = [
            'bugged', 'don\'t use', 'dont use', 'broken', 'removed', 
            'disabled', 'corrupt', 'error', 'invalid', 'test item',
            'debug', 'do not use', 'not working'
        ]
        
        if notes_col and notes_col in df.columns:
            for idx, row in df.iterrows():
                notes = str(row[notes_col]).lower() if pd.notna(row[notes_col]) else ""
                if any(indicator in notes for indicator in problematic_indicators):
                    problematic_rows.add(idx)
        
        # Also filter rows where item name contains problematic indicators
        name_col = self._find_column(df, ['name', 'itemname', 'item_name', 'item name'])
        if name_col:
            for idx, row in df.iterrows():
                name = str(row[name_col]).lower() if pd.notna(row[name_col]) else ""
                if any(indicator in name for indicator in problematic_indicators):
                    problematic_rows.add(idx)
        
        logger.info(f"Found {len(problematic_rows)} problematic rows to filter out (text-based)")
        return problematic_rows
    
    def _map_category_string(self, category_str: str) -> Optional[ItemCategory]:
        """Map a category string from Excel to ItemCategory enum"""
        if not category_str:
            return None
        
        category_lower = category_str.lower().strip()
        
        # Skip invalid/placeholder categories
        invalid_categories = ['nothing', 'dont use', 'don\'t use', '', 'n/a', 'na', 'null', 'none', 'debug item', 'debug', 'loot table']
        if category_lower in invalid_categories:
            return None
        
        # Direct mappings for DDV categories
        mappings = {
            # Furniture/Decor (largest category in your Excel)
            'furniture': ItemCategory.FURNITURE,
            'decor': ItemCategory.FURNITURE,
            'decoration': ItemCategory.FURNITURE,
            'furnishing': ItemCategory.FURNITURE,
            'home': ItemCategory.FURNITURE,
            
            # Clothing subcategories
            'clothes outfit': ItemCategory.CLOTHES_OUTFITS,
            'clothes jacket': ItemCategory.CLOTHES_TOPS,
            'clothes top': ItemCategory.CLOTHES_TOPS,
            'clothes bottom': ItemCategory.CLOTHES_BOTTOMS,
            'clothes skirt': ItemCategory.CLOTHES_BOTTOMS,
            'clothes helmet': ItemCategory.CLOTHES_HELMETS,
            'clothes hair': ItemCategory.CLOTHES_HELMETS,
            'clothes shoe': ItemCategory.CLOTHES_SHOES,
            'clothes socks': ItemCategory.CLOTHES_SHOES,
            
            # Clothing accessories
            'clothes back': ItemCategory.CLOTHES_ACCESSORIES,
            'clothes glove': ItemCategory.CLOTHES_ACCESSORIES,
            'clothes glasses': ItemCategory.CLOTHES_ACCESSORIES,
            'clothes bracelet': ItemCategory.CLOTHES_ACCESSORIES,
            'clothes earring': ItemCategory.CLOTHES_ACCESSORIES,
            'clothes inner neckwear': ItemCategory.CLOTHES_ACCESSORIES,
            
            # General clothing terms (fallback)
            'clothing': ItemCategory.CLOTHES_OTHER,
            'clothes': ItemCategory.CLOTHES_OTHER,
            'apparel': ItemCategory.CLOTHES_OTHER,
            'fashion': ItemCategory.CLOTHES_OTHER,
            'motif': ItemCategory.CLOTHES_OTHER,  # DDV motifs are clothing patterns
            
            # Food
            'food': ItemCategory.FOOD,
            'meal': ItemCategory.FOOD,
            'ingredient': ItemCategory.FOOD,
            'recipe': ItemCategory.FOOD,
            
            # Tools
            'tool': ItemCategory.TOOLS,
            'tools': ItemCategory.TOOLS,
            'equipment': ItemCategory.TOOLS,
            'fishing rod': ItemCategory.TOOLS,
            'shovel': ItemCategory.TOOLS,
            'pickaxe': ItemCategory.TOOLS,
            'watering can': ItemCategory.TOOLS,
            
            # Materials
            'material': ItemCategory.MATERIALS,
            'materials': ItemCategory.MATERIALS,
            'resource': ItemCategory.MATERIALS,
            'gem': ItemCategory.MATERIALS,
            'ore': ItemCategory.MATERIALS,
            'stone': ItemCategory.MATERIALS,
            'wood': ItemCategory.MATERIALS,
            'flower': ItemCategory.MATERIALS,
            
            # Pets/Companions
            'pet': ItemCategory.PETS,
            'pets': ItemCategory.PETS,
            'companion': ItemCategory.PETS,
            'critter': ItemCategory.PETS,
            
            # House subcategories
            'house': ItemCategory.HOUSE_SKINS,
            'houses': ItemCategory.HOUSE_SKINS,
            'building': ItemCategory.HOUSE_SKINS,
            'structure': ItemCategory.HOUSE_SKINS,
            'house wallpaper': ItemCategory.HOUSE_WALLPAPER,
            'house floor': ItemCategory.HOUSE_FLOORS,
            'npc house': ItemCategory.NPC_HOUSES,
            
            # Character skins
            'skin': ItemCategory.NPC_SKINS,
            'skins': ItemCategory.NPC_SKINS,
            'character': ItemCategory.NPC_SKINS,
            'npc model': ItemCategory.NPC_SKINS,
            'npc': ItemCategory.NPC_SKINS,
        }
        
        # Check exact matches first
        if category_lower in mappings:
            return mappings[category_lower]
        
        # Check partial matches
        for key, category in mappings.items():
            if key in category_lower or category_lower in key:
                return category
        
        # Smart fallbacks based on category content
        if 'clothes' in category_lower or 'outfit' in category_lower:
            logger.debug(f"Unknown clothing category '{category_str}', defaulting to clothes_other")
            return ItemCategory.CLOTHES_OTHER
        elif 'house' in category_lower or 'building' in category_lower:
            logger.debug(f"Unknown house category '{category_str}', defaulting to house_skins")
            return ItemCategory.HOUSE_SKINS
        elif 'furniture' in category_lower or 'decor' in category_lower:
            logger.debug(f"Unknown furniture category '{category_str}', defaulting to furniture")
            return ItemCategory.FURNITURE
        
        # Final fallback
        logger.debug(f"Completely unknown category '{category_str}', skipping")
        return None
    
    def _process_worksheet(self, df: pd.DataFrame, category: ItemCategory, sheet_name: str) -> List[GameItem]:
        """Process a single worksheet into GameItem objects"""
        items = []
        
        if df.empty:
            logger.warning(f"Empty worksheet: {sheet_name}")
            return items
        
        # Normalize column names
        df.columns = [col.strip().lower() if isinstance(col, str) else str(col) for col in df.columns]
        
        # Find required columns
        id_col = self._find_column(df, ['id', 'itemid', 'item_id', 'petitemid', 'clothesitemid', 'houseitemid'])
        name_col = self._find_column(df, ['name', 'itemname', 'item_name', 'namepets', 'clothesname', 'housename'])
        
        if not id_col:
            logger.error(f"No ID column found in sheet: {sheet_name}")
            return items
        
        if not name_col:
            logger.error(f"No name column found in sheet: {sheet_name}")
            return items
        
        # Find optional columns
        desc_col = self._find_column(df, ['description', 'desc', 'notes', 'info'])
        image_col = self._find_column(df, ['image', 'imagepath', 'icon', 'picture'])
        rarity_col = self._find_column(df, ['rarity', 'tier', 'grade'])
        cost_col = self._find_column(df, ['cost', 'price', 'value'])
        
        # Process each row
        for idx, row in df.iterrows():
            try:
                # Get ID and validate
                item_id = row[id_col]
                if pd.isna(item_id) or not isinstance(item_id, (int, float)) or item_id <= 0:
                    continue
                
                item_id = int(item_id)
                
                # Get name and validate
                item_name = row[name_col]
                if pd.isna(item_name) or not str(item_name).strip():
                    item_name = f"Item {item_id}"
                
                item_name = str(item_name).strip()
                
                # Get optional fields
                description = None
                if desc_col and not pd.isna(row[desc_col]):
                    description = str(row[desc_col]).strip()
                
                image_path = None
                if image_col and not pd.isna(row[image_col]):
                    image_path = str(row[image_col]).strip()
                
                rarity = None
                if rarity_col and not pd.isna(row[rarity_col]):
                    rarity = str(row[rarity_col]).strip()
                
                cost = None
                if cost_col and not pd.isna(row[cost_col]):
                    try:
                        cost = int(float(row[cost_col]))
                    except (ValueError, TypeError):
                        pass
                
                # Collect custom data from other columns
                custom_data = {}
                for col in df.columns:
                    if col not in [id_col, name_col, desc_col, image_col, rarity_col, cost_col]:
                        value = row[col]
                        if not pd.isna(value):
                            custom_data[col] = value
                
                # Create GameItem
                item = GameItem(
                    id=item_id,
                    name=item_name,
                    category=category,
                    description=description,
                    image_path=image_path,
                    rarity=rarity,
                    cost=cost,
                    custom_data=custom_data
                )
                
                items.append(item)
                
            except Exception as e:
                logger.warning(f"Error processing row {idx} in {sheet_name}: {e}")
                continue
        
        return items
    
    def _find_column(self, df: pd.DataFrame, possible_names: List[str]) -> Optional[str]:
        """Find a column by trying multiple possible names"""
        df_columns_lower = [col.lower() for col in df.columns]
        
        for name in possible_names:
            name_lower = name.lower()
            if name_lower in df_columns_lower:
                # Return the original column name (with original case)
                idx = df_columns_lower.index(name_lower)
                return df.columns[idx]
        
        return None
    
    def get_available_sheets(self) -> List[str]:
        """Get list of available worksheets in the Excel file"""
        try:
            if not self.excel_path.exists():
                return []
            
            excel_file = pd.ExcelFile(self.excel_path, engine='openpyxl')
            return excel_file.sheet_names
        except Exception as e:
            logger.error(f"Error reading Excel sheets: {e}")
            return []
    
    def validate_excel_file(self) -> Tuple[bool, List[str]]:
        """Validate the Excel file and return status and issues"""
        issues = []
        
        try:
            if not self.excel_path.exists():
                issues.append(f"File not found: {self.excel_path}")
                return False, issues
            
            # Try to read the file
            excel_data = pd.read_excel(self.excel_path, sheet_name=None, engine='openpyxl')
            
            if not excel_data:
                issues.append("No worksheets found in Excel file")
                return False, issues
            
            valid_sheets = 0
            for sheet_name, df in excel_data.items():
                category = self._detect_category(sheet_name)
                if category:
                    # Check for required columns
                    id_col = self._find_column(df, ['id', 'itemid', 'item_id'])
                    name_col = self._find_column(df, ['name', 'itemname', 'item_name'])
                    
                    if not id_col:
                        issues.append(f"Sheet '{sheet_name}': No ID column found")
                    elif not name_col:
                        issues.append(f"Sheet '{sheet_name}': No name column found")
                    else:
                        valid_sheets += 1
            
            if valid_sheets == 0:
                issues.append("No valid worksheets found with required columns")
                return False, issues
            
            return True, issues
            
        except Exception as e:
            issues.append(f"Error reading Excel file: {e}")
            return False, issues
    
    def export_to_excel(self, database: GameDatabase, output_path: str) -> bool:
        """Export game database back to Excel format"""
        try:
            output_path = Path(output_path)
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for category, collection in database.collections.items():
                    if len(collection) == 0:
                        continue
                    
                    # Prepare data for DataFrame
                    data = []
                    for item in collection:
                        row = {
                            'ID': item.id,
                            'Name': item.name,
                            'Description': item.description or '',
                            'Image': item.image_path or '',
                            'Rarity': item.rarity or '',
                            'Cost': item.cost or ''
                        }
                        # Add custom data
                        for key, value in item.custom_data.items():
                            row[key.title()] = value
                        
                        data.append(row)
                    
                    # Create DataFrame and save to sheet
                    df = pd.DataFrame(data)
                    sheet_name = category.value.title()
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            logger.info(f"Database exported to: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            return False
